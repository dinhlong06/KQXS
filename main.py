import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import win32com.client
import time
import os
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import Toplevel

file_path = "example.xlsx"

# Cấu hình tỉnh theo ngày và tên hiển thị
PROVINCE_CONFIG = {
    "urls": {
        0: ["tp-hcm", "dong-thap", "ca-mau"],
        1: ["ben-tre", "vung-tau", "bac-lieu"],
        2: ["dong-nai", "can-tho", "soc-trang"],
        3: ["tay-ninh", "an-giang", "binh-thuan"],
        4: ["vinh-long", "binh-duong", "tra-vinh"],
        5: ["tp-hcm", "long-an", "binh-phuoc"],
        6: ["tien-giang", "kien-giang", "da-lat"]
    },
    "display_names": [
        ["TP.HCM", "Đ.THÁP", "CÀ MAU"],
        ["BẾN TRE", "V.TÀU", "BẠC LIÊU"],
        ["ĐỒNG NAI", "CẦN THƠ", "S.TRĂNG"],
        ["TÂY NINH", "AN GIANG", "B.THUẬN"],
        ["V.LONG", "B.DƯƠNG", "T.VINH"],
        ["TP.HCM", "LONG AN", "B.PHƯỚC"],
        ["T.GIANG", "K.GIANG", "ĐÀ LẠT"]
    ]
}

def close_excel_file(file_path):
    """Đóng file Excel nếu nó đang mở."""
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for workbook in excel.Workbooks:
            if workbook.FullName == os.path.abspath(file_path):
                workbook.Close(SaveChanges=False)
        excel.Quit()
        return True
    except Exception as e:
        print(f"Không thể đóng file Excel: {e}")
        return False

def open_excel_file(file_path):
    """Mở lại file Excel sau khi hoàn tất."""
    try:
        os.startfile(file_path)
    except Exception as e:
        print(f"Không thể mở lại file: {e}")

def province(thu):
    """Lấy danh sách tỉnh theo ngày trong tuần."""
    return PROVINCE_CONFIG["urls"].get(thu, [])

def fetch_url(url, ngayxoso):
    """Truy cập URL và xử lý dữ liệu."""
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return xuly(response, ngayxoso)
    except requests.RequestException as e:
        print(f"Lỗi khi truy cập {url}: {e}")
        return {}

def xuly(response, ngayxoso):
    """Xử lý dữ liệu HTML để lấy kết quả xổ số."""
    try:
        soup = BeautifulSoup(response.text, "html.parser")
        ngay_elements = soup.find_all("div", class_="ngay")
        if not ngay_elements:
            print("Không tìm thấy thông tin ngày xổ số.")
            return {}

        found_date = False
        for x in ngay_elements:
            for a in x.find_all("a"):
                if a.get_text(strip=True) == ngayxoso:
                    found_date = True
                    break
            if found_date:
                break

        if not found_date:
            print(f"Ngày {ngayxoso} không khớp với dữ liệu trên trang. Chưa xổ đài này!")
            return {}

        giais = soup.find("table", class_="box_kqxs_content")
        if not giais:
            print("Không tìm thấy bảng kết quả xổ số.")
            return {}

        tinh = {
            "giaidb": [], "giai1": [], "giai2": [], "giai3": [],
            "giai4": [], "giai5": [], "giai6": [], "giai7": [], "giai8": []
        }

        keys = set(tinh.keys())
        for td in giais.find_all("td"):
            tdclass = td.get("class", [])
            if tdclass and tdclass[0] in keys:
                for div in td.find_all("div"):
                    tinh[tdclass[0]].append(div.get_text(strip=True))

        return dict(reversed(tinh.items()))
    except Exception as e:
        print(f"Lỗi khi xử lý dữ liệu từ {response.url}: {e}")
        return {}

def write_to_excel(tinhs, weekday, ngayxoso):
    """Ghi dữ liệu vào file Excel."""
    try:
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        sheet = wb.active

        num_provinces = len(PROVINCE_CONFIG["display_names"][weekday])
        for row in range(2, 20 + 2):
            for col in range(2, num_provinces + 2):
                sheet.cell(row, col).value = None

        for i in range(num_provinces):
            sheet.cell(1, i + 2).value = PROVINCE_CONFIG["display_names"][weekday][i]

        for i, tinh in enumerate(tinhs):
            if tinh:
                row = 2
                for key, values in tinh.items():
                    for value in values:
                        sheet.cell(row, i + 2).value = value
                        row += 1

        sheet.cell(24, 1).value = "KQXS NGÀY " + ngayxoso
        wb.save(file_path)
        wb.close()
        print("Ghi dữ liệu vào Excel thành công.")
    except Exception as e:
        print(f"Lỗi khi ghi vào Excel: {e}")

# Biến toàn cục để kiểm soát việc cập nhật thời gian thực
running = False
task_id = None
last_update_date = None

def is_data_complete(tinhs):
    """Kiểm tra xem dữ liệu đã đầy đủ (có kết quả cho tất cả tỉnh) chưa."""
    # Nếu bất kỳ tỉnh nào không có dữ liệu (tức là dictionary rỗng), trả về False
    return all(tinh and all(tinh.get(key) for key in tinh) for tinh in tinhs)

def main(day_ago=0, show_message=True):
    """Hàm chính để chạy chương trình. show_message quyết định có hiển thị thông báo không."""
    global last_update_date
    now = datetime.datetime.today() - datetime.timedelta(days=day_ago)
    weekday = now.weekday()
    ngayxoso = str(now.strftime('%d/%m/%Y'))
    last_update_date = ngayxoso

    urls = [f"https://www.minhngoc.net/xo-so-mien-nam/{tinh}.html" for tinh in PROVINCE_CONFIG["urls"][weekday]]

    with ThreadPoolExecutor(max_workers=3) as executor:
        tinhs = list(executor.map(fetch_url, urls, [ngayxoso] * len(urls)))

    if os.path.exists(file_path):
        print("Đang kiểm tra file Excel...")
        if close_excel_file(file_path):
            time.sleep(1)
        else:
            messagebox.showerror("Lỗi", "Không thể đóng file Excel. Vui lòng đóng thủ công và thử lại.")
            return False

    write_to_excel(tinhs, weekday, ngayxoso)
    if show_message:
        messagebox.showinfo("Thông báo", f"Đã thực hiện cho ngày {ngayxoso}")
    open_excel_file(file_path)
    return True

def run_once(day_ago):
    """Chạy một lần duy nhất."""
    main(day_ago, show_message=True)

def run_realtime(top, status_label):
    """Chạy theo thời gian thực mỗi 5 giây, với nhãn trạng thái trong cửa sổ con."""
    global running, task_id
    if running:
        status_label.config(text="Đang cập nhật thời gian thực...")
        now = datetime.datetime.today()
        weekday = now.weekday()
        ngayxoso = str(now.strftime('%d/%m/%Y'))

        urls = [f"https://www.minhngoc.net/xo-so-mien-nam/{tinh}.html" for tinh in PROVINCE_CONFIG["urls"][weekday]]
        with ThreadPoolExecutor(max_workers=3) as executor:
            tinhs = list(executor.map(fetch_url, urls, [ngayxoso] * len(urls)))

        if is_data_complete(tinhs):
            stop_realtime(top, status_label)
            messagebox.showinfo("Thông báo", f"Dữ liệu đã đầy đủ cho ngày {ngayxoso}. Đã dừng cập nhật.")
            open_excel_file(file_path)
        elif main(0, show_message=False):
            task_id = top.after(5000, lambda: run_realtime(top, status_label))

def stop_realtime(top, status_label):
    """Dừng cập nhật thời gian thực."""
    global running, task_id
    running = False
    if task_id is not None:
        top.after_cancel(task_id)
        task_id = None
    status_label.config(text="Đã dừng cập nhật.")
    if last_update_date:
        messagebox.showinfo("Thông báo", f"Cập nhật thời gian thực đã dừng.\nLần cập nhật cuối: {last_update_date}")

def show_today_options():
    """Hiển thị cửa sổ con để chọn chế độ cho Hôm nay."""
    global running, task_id
    running = False  # Đảm bảo dừng nếu đang chạy

    top = Toplevel(root)
    top.title("Chọn chế độ Hôm nay")
    top.geometry("300x300")
    top.configure(bg="#f0f0f0")

    label = ttk.Label(top, text="Chọn chế độ", font=("Helvetica", 14, "bold"), background="#f0f0f0")
    label.pack(pady=10)

    btn_once = ttk.Button(top, text="Cập nhật một lần", command=lambda: [run_once(0), top.destroy()], style="TButton")
    btn_once.pack(pady=5)

    status_label = ttk.Label(top, text="", font=("Helvetica", 12), background="#f0f0f0", foreground="#666")
    status_label.pack(pady=5)

    def start_realtime():
        global running
        running = True
        btn_stop.pack(pady=5)  # Hiển thị nút dừng trong cửa sổ con
        run_realtime(top, status_label)

    btn_realtime = ttk.Button(top, text="Cập nhật thời gian thực", command=start_realtime, style="TButton")
    btn_realtime.pack(pady=5)

    btn_stop = ttk.Button(top, text="Dừng", command=lambda: [stop_realtime(top, status_label), top.destroy()], style="TButton")
    # Nút dừng ban đầu không hiển thị, chỉ hiển thị khi chạy thời gian thực

# Tạo giao diện GUI
root = tk.Tk()
root.title("Kết Quả Xổ Số")
root.geometry("500x300")
root.configure(bg="#f0f0f0")

title_label = ttk.Label(root, text="Kết Quả Xổ Số", font=("Helvetica", 20, "bold"), background="#f0f0f0", foreground="#333")
title_label.pack(pady=20)

button_frame = ttk.Frame(root, style="TFrame")
button_frame.pack(pady=10)

style = ttk.Style()
style.configure("TButton", font=("Helvetica", 12), padding=10)

btn_today = ttk.Button(button_frame, text="Hôm Nay", command=show_today_options, style="TButton")
btn_today.grid(row=0, column=0, padx=10, pady=10)

btn_yesterday = ttk.Button(button_frame, text="Hôm Qua", command=lambda: run_once(1), style="TButton")
btn_yesterday.grid(row=0, column=1, padx=10, pady=10)

btn_two_days_ago = ttk.Button(button_frame, text="2 Ngày Trước", command=lambda: run_once(2), style="TButton")
btn_two_days_ago.grid(row=0, column=2, padx=10, pady=10)

if __name__ == "__main__":
    root.mainloop()